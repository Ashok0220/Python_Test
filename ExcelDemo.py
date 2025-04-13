

import openpyxl

book = openpyxl.load_workbook("C:\\Users\\ashok\\OneDrive\\Desktop\\Pythonexcel.xlsx")

Dics = { }
# Control to Active the sheet
sheet = book.active

cell = sheet.cell(row=1, column=2)
print(cell.value)

sheet.cell(row=2, column=2).value = "Ashok"

print(sheet.cell(row=2, column=2).value)

print(sheet.max_row)

print(sheet.max_column)

# another way
print(sheet['A4'].value)

# get the 1st colum value only

# for i in range(1, sheet.max_row + 1):
#     print(sheet.cell(row=i, column=1).value)


# get all row and colums
for i in range(1, sheet.max_row + 1): # to get row
    if sheet.cell(row=i, column=1).value == "TestCase3":  # if i want testcase3 all data only
        for j in range(1, sheet.max_column + 1): # to get colum
            Dics[sheet.cell(row=1, column=j).value] = sheet.cell(row=i,column=j).value


print(Dics)




